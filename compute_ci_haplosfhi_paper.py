import csv
import openpyxl

from aux import bootstrap_simulation


def haplosfhi_bootstrap_simulation() -> None:
    # ### Accuracies
    # print("Accuracies")
    # ## HAPLOSFHI
    # print("HaploSFHI")
    # document_c1 = openpyxl.load_workbook(
    #     r"C:\Users\lhott\Documents\Formation scolaire\Stages\Stage de fin d'étude\Analyse"
    #     r"\NotreExtrapo\Version_Finale\Classe I\Tests Classe I\Archive\ResultatsNotreExtrapoDefCI.xlsx",
    #     data_only=True,
    # )
    # feuille_c1 = document_c1["Score2_Allèles"]
    #
    # document_2145 = openpyxl.load_workbook(
    #     r"C:\Users\lhott\Documents\Formation scolaire"
    #     r"\These\Travail\SagaGL\HaploSFHI\HaploSFHI Test Set\scoresSFHI.xlsx",
    #     data_only=True,
    # )
    # feuille_2145 = document_2145["2digits complet"]
    #
    # document_drb345 = openpyxl.load_workbook(
    #     r"C:\Users\lhott\Documents\Formation scolaire\Stages\Stage de fin d'étude"
    #     r"\Analyse\NotreExtrapo\Version_Finale\Classe II\DRB345\Tests Classe II\ResultatsNotreExtrapoDef 83.xlsx",
    #     data_only=True,
    # )
    # feuille_drb345 = document_drb345["Feuil1 amélioration"]
    #
    # document_dqa1 = openpyxl.load_workbook(
    #     r"C:\Users\lhott\Documents\Formation scolaire\Stages\Stage de fin d'étude\Analyse"
    #     r"\NotreExtrapo\Version_Finale\Classe II\DQA\Tests Classe II\ResultatsNotreExtrapoDefDQA.xlsx",
    #     data_only=True,
    # )
    # feuille_dqa1 = document_dqa1["Feuil1"]
    #
    # # A
    # haplosfhi_results_list_a: list[float] = [
    #     100*feuille_c1.cell(row=row_nu, column=15).value/2 for row_nu in range(2, 3886)
    # ]
    # x_low, x_high = bootstrap_simulation(haplosfhi_results_list_a, 25000, len(haplosfhi_results_list_a))
    # print(f"Bootstrap CI estimates HLA-A: {round(x_low, 2)} {round(x_high, 2)}")
    # # B
    # haplosfhi_results_list_b: list[float] = [
    #     100*feuille_c1.cell(row=row_nu, column=17).value/2 for row_nu in range(2, 3886)
    # ]
    # x_low, x_high = bootstrap_simulation(haplosfhi_results_list_b, 25000, len(haplosfhi_results_list_b))
    # print(f"Bootstrap CI estimates HLA-B: {round(x_low, 2)} {round(x_high, 2)}")
    # # C
    # haplosfhi_results_list_c: list[float] = [
    #     100*feuille_c1.cell(row=row_nu, column=19).value/2 for row_nu in range(2, 3886)
    # ]
    # x_low, x_high = bootstrap_simulation(haplosfhi_results_list_c, 25000, len(haplosfhi_results_list_c))
    # print(f"Bootstrap CI estimates HLA-C: {round(x_low, 2)} {round(x_high, 2)}")
    #
    # # DRB1
    # haplosfhi_results_list_drb1: list[float] = [
    #     100*feuille_2145.cell(row=row_nu, column=29).value/2 for row_nu in range(3, 2148)
    # ]
    # x_low, x_high = bootstrap_simulation(haplosfhi_results_list_drb1, 25000, len(haplosfhi_results_list_drb1))
    # print(f"Bootstrap CI estimates HLA-DRB1: {round(x_low, 2)} {round(x_high, 2)}")
    # # DQB1
    # haplosfhi_results_list_dqb1: list[float] = [
    #     100*feuille_2145.cell(row=row_nu, column=31).value/2 for row_nu in range(3, 2148)
    # ]
    # x_low, x_high = bootstrap_simulation(haplosfhi_results_list_dqb1, 25000, len(haplosfhi_results_list_dqb1))
    # print(f"Bootstrap CI estimates HLA-DQB1: {round(x_low, 2)} {round(x_high, 2)}")
    #
    # # DRB345
    # haplosfhi_results_list_drb345: list[float] = [
    #     100*feuille_drb345.cell(row=row_nu, column=19).value/2 for row_nu in range(2, 905)
    # ]
    # x_low, x_high = bootstrap_simulation(haplosfhi_results_list_drb345, 25000, len(haplosfhi_results_list_drb345))
    # print(f"Bootstrap CI estimates HLA-DRB345: {round(x_low, 2)} {round(x_high, 2)}")
    #
    # # DQA1
    # haplosfhi_results_list_dqa1: list[float] = [
    #     100*feuille_dqa1.cell(row=row_nu, column=19).value/2 for row_nu in range(2, 2433)
    # ]
    # x_low, x_high = bootstrap_simulation(haplosfhi_results_list_dqa1, 25000, len(haplosfhi_results_list_dqa1))
    # print(f"Bootstrap CI estimates HLA-DQA1: {round(x_low, 2)} {round(x_high, 2)}")
    #
    #
    # ## HaploStats (CAU)
    # print("HaploStats (CAU)")
    # document_haplostats_normal = openpyxl.load_workbook(
    #     r"C:\Users\lhott\Documents\Formation scolaire\These\Travail\SagaGL"
    #     r"\HaploSFHI\HaploStats Test Set\scoresCAU.xlsx",
    #     data_only=True,
    # )
    # feuille_haplostats_normal = document_haplostats_normal["2digits complet"]
    # document_haplostats_drb345 = openpyxl.load_workbook(
    #     r"C:\Users\lhott\Documents\Formation scolaire\These\Travail\SagaGL"
    #     r"\HaploSFHI\HaploStats Test Set\HaploStats Test DRB345\scores.xlsx",
    #     data_only=True,
    # )
    # feuille_haplostats_drb345 = document_haplostats_drb345["Feuil1"]
    #
    # # A
    # haplostats_cau_results_a: list[float] = [
    #     100 * feuille_haplostats_normal.cell(row=row_nu, column=23).value / 2 for row_nu in range(3, 2148)
    # ]
    # x_low, x_high = bootstrap_simulation(haplostats_cau_results_a, 25000, len(haplostats_cau_results_a))
    # print(f"Bootstrap CI estimates HLA-A: {round(x_low, 2)} {round(x_high, 2)}")
    # # B
    # haplostats_cau_results_b: list[float] = [
    #     100 * feuille_haplostats_normal.cell(row=row_nu, column=25).value / 2 for row_nu in range(3, 2148)
    # ]
    # x_low, x_high = bootstrap_simulation(haplostats_cau_results_b, 25000, len(haplostats_cau_results_b))
    # print(f"Bootstrap CI estimates HLA-B: {round(x_low, 2)} {round(x_high, 2)}")
    # # C
    # haplostats_cau_results_c: list[float] = [
    #     100 * feuille_haplostats_normal.cell(row=row_nu, column=27).value / 2 for row_nu in range(3, 2148)
    # ]
    # x_low, x_high = bootstrap_simulation(haplostats_cau_results_c, 25000, len(haplostats_cau_results_c))
    # print(f"Bootstrap CI estimates HLA-C: {round(x_low, 2)} {round(x_high, 2)}")
    # # DRB1
    # haplostats_cau_results_drb1: list[float] = [
    #     100 * feuille_haplostats_normal.cell(row=row_nu, column=29).value / 2 for row_nu in range(3, 2148)
    # ]
    # x_low, x_high = bootstrap_simulation(haplostats_cau_results_drb1, 25000, len(haplostats_cau_results_drb1))
    # print(f"Bootstrap CI estimates HLA-DRB1: {round(x_low, 2)} {round(x_high, 2)}")
    # # DQB1
    # haplostats_cau_results_dqb1: list[float] = [
    #     100 * feuille_haplostats_normal.cell(row=row_nu, column=31).value / 2 for row_nu in range(3, 2148)
    # ]
    # x_low, x_high = bootstrap_simulation(haplostats_cau_results_dqb1, 25000, len(haplostats_cau_results_dqb1))
    # print(f"Bootstrap CI estimates HLA-DQB1: {round(x_low, 2)} {round(x_high, 2)}")
    #
    # # DRB345
    # haplostats_cau_results_drb345: list[float] = [
    #     100 * feuille_haplostats_drb345.cell(row=row_nu, column=37).value / 2 for row_nu in range(3, 1287)
    # ]
    # x_low, x_high = bootstrap_simulation(haplostats_cau_results_drb345, 25000, len(haplostats_cau_results_drb345))
    # print(f"Bootstrap CI estimates HLA-DRB345: {round(x_low, 2)} {round(x_high, 2)}")
    #
    # ## HaploStats (EURCAU)
    # print("HaploStats (EURCAU)")
    # document_haplostats_normal = openpyxl.load_workbook(
    #     r"C:\Users\lhott\Documents\Formation scolaire\These\Travail\SagaGL"
    #     r"\HaploSFHI\HaploStats Test Set\HaploStats Test EURCAU\scoresEURCAU.xlsx",
    #     data_only=True,
    # )
    # feuille_haplostats_normal = document_haplostats_normal["2digits complet"]
    # document_haplostats_drb345 = openpyxl.load_workbook(
    #     r"C:\Users\lhott\Documents\Formation scolaire\These\Travail\SagaGL"
    #     r"\HaploSFHI\HaploStats Test Set\HaploStats Test DRB345 - EURCAU\scores.xlsx",
    #     data_only=True,
    # )
    # feuille_haplostats_drb345 = document_haplostats_drb345["Feuil1"]
    #
    # # A
    # haplostats_cau_results_a: list[float] = [
    #     100 * feuille_haplostats_normal.cell(row=row_nu, column=23).value / 2 for row_nu in range(3, 2148)
    # ]
    # x_low, x_high = bootstrap_simulation(haplostats_cau_results_a, 25000, len(haplostats_cau_results_a))
    # print(f"Bootstrap CI estimates HLA-A: {round(x_low, 2)} {round(x_high, 2)}")
    # # B
    # haplostats_cau_results_b: list[float] = [
    #     100 * feuille_haplostats_normal.cell(row=row_nu, column=25).value / 2 for row_nu in range(3, 2148)
    # ]
    # x_low, x_high = bootstrap_simulation(haplostats_cau_results_b, 25000, len(haplostats_cau_results_b))
    # print(f"Bootstrap CI estimates HLA-B: {round(x_low, 2)} {round(x_high, 2)}")
    # # C
    # haplostats_cau_results_c: list[float] = [
    #     100 * feuille_haplostats_normal.cell(row=row_nu, column=27).value / 2 for row_nu in range(3, 2148)
    # ]
    # x_low, x_high = bootstrap_simulation(haplostats_cau_results_c, 25000, len(haplostats_cau_results_c))
    # print(f"Bootstrap CI estimates HLA-C: {round(x_low, 2)} {round(x_high, 2)}")
    # # DRB1
    # haplostats_cau_results_drb1: list[float] = [
    #     100 * feuille_haplostats_normal.cell(row=row_nu, column=29).value / 2 for row_nu in range(3, 2148)
    # ]
    # x_low, x_high = bootstrap_simulation(haplostats_cau_results_drb1, 25000, len(haplostats_cau_results_drb1))
    # print(f"Bootstrap CI estimates HLA-DRB1: {round(x_low, 2)} {round(x_high, 2)}")
    # # DQB1
    # haplostats_cau_results_dqb1: list[float] = [
    #     100 * feuille_haplostats_normal.cell(row=row_nu, column=31).value / 2 for row_nu in range(3, 2148)
    # ]
    # x_low, x_high = bootstrap_simulation(haplostats_cau_results_dqb1, 25000, len(haplostats_cau_results_dqb1))
    # print(f"Bootstrap CI estimates HLA-DQB1: {round(x_low, 2)} {round(x_high, 2)}")
    #
    # # DRB345
    # haplostats_cau_results_drb345: list[float] = [
    #     100 * feuille_haplostats_drb345.cell(row=row_nu, column=37).value / 2 for row_nu in range(3, 1287)
    # ]
    # x_low, x_high = bootstrap_simulation(haplostats_cau_results_drb345, 25000, len(haplostats_cau_results_drb345))
    # print(f"Bootstrap CI estimates HLA-DRB345: {round(x_low, 2)} {round(x_high, 2)}")
    #
    # ## HLA-EMMA
    # print("HLA-EMMA")
    # document_hla_emma = openpyxl.load_workbook(
    #     r"C:\Users\lhott\Documents\Formation scolaire\These\Travail\SagaGL\HaploSFHI"
    #     r"\virtual HLA-EMMA for accuracy map\scores.xlsx",
    #     data_only=True,
    # )
    # feuille_hla_emma = document_hla_emma["Score2_Allèles"]
    #
    # # A
    # hlaemma_results_a: list[float] = [
    #     100 * feuille_hla_emma.cell(row=row_nu, column=23).value / 2 for row_nu in range(3, 2148)
    # ]
    # x_low, x_high = bootstrap_simulation(hlaemma_results_a, 25000, len(hlaemma_results_a))
    # print(f"Bootstrap CI estimates HLA-A: {round(x_low, 2)} {round(x_high, 2)}")
    # # B
    # hlaemma_results_b: list[float] = [
    #     100 * feuille_hla_emma.cell(row=row_nu, column=25).value / 2 for row_nu in range(3, 2433)
    # ]
    # x_low, x_high = bootstrap_simulation(hlaemma_results_b, 25000, len(hlaemma_results_b))
    # print(f"Bootstrap CI estimates HLA-B: {round(x_low, 2)} {round(x_high, 2)}")
    # # C
    # hlaemma_results_c: list[float] = [
    #     100 * feuille_hla_emma.cell(row=row_nu, column=27).value / 2 for row_nu in range(3, 2433)
    # ]
    # x_low, x_high = bootstrap_simulation(hlaemma_results_c, 25000, len(hlaemma_results_c))
    # print(f"Bootstrap CI estimates HLA-C: {round(x_low, 2)} {round(x_high, 2)}")
    # # DRB1
    # hlaemma_results_drb1: list[float] = [
    #     100 * feuille_hla_emma.cell(row=row_nu, column=29).value / 2 for row_nu in range(3, 2433)
    # ]
    # x_low, x_high = bootstrap_simulation(hlaemma_results_drb1, 25000, len(hlaemma_results_drb1))
    # print(f"Bootstrap CI estimates HLA-DRB1: {round(x_low, 2)} {round(x_high, 2)}")
    # # DQB1
    # hlaemma_results_dqb1: list[float] = [
    #     100 * feuille_hla_emma.cell(row=row_nu, column=31).value / 2 for row_nu in range(3, 2433)
    # ]
    # x_low, x_high = bootstrap_simulation(hlaemma_results_dqb1, 25000, len(hlaemma_results_dqb1))
    # print(f"Bootstrap CI estimates HLA-DQB1: {round(x_low, 2)} {round(x_high, 2)}")
    #
    # ## HLA-Upgrade (FR)
    # print("HLA-Upgrade (FR)")
    # document_hla_upgrade = openpyxl.load_workbook(
    #     r"C:\Users\lhott\Documents\Formation scolaire\These\Travail\SagaGL\HaploSFHI"
    #     r"\easyHLA Test Set\scoresFR.xlsx",
    #     data_only=True,
    # )
    # feuille_hla_upgrade = document_hla_upgrade["Feuil1"]
    #
    # # A
    # hlaupgrade_results_a: list[float] = [
    #     100 * feuille_hla_upgrade.cell(row=row_nu, column=24).value / 2 for row_nu in range(3, 2148)
    # ]
    # x_low, x_high = bootstrap_simulation(hlaupgrade_results_a, 25000, len(hlaupgrade_results_a))
    # print(f"Bootstrap CI estimates HLA-A: {round(x_low, 2)} {round(x_high, 2)}")
    # # B
    # hlaupgrade_results_b: list[float] = [
    #     100 * feuille_hla_upgrade.cell(row=row_nu, column=26).value / 2 for row_nu in range(3, 2148)
    # ]
    # x_low, x_high = bootstrap_simulation(hlaupgrade_results_b, 25000, len(hlaupgrade_results_b))
    # print(f"Bootstrap CI estimates HLA-B: {round(x_low, 2)} {round(x_high, 2)}")
    # # C
    # hlaupgrade_results_c: list[float] = [
    #     100 * feuille_hla_upgrade.cell(row=row_nu, column=28).value / 2 for row_nu in range(3, 2148)
    # ]
    # x_low, x_high = bootstrap_simulation(hlaupgrade_results_c, 25000, len(hlaupgrade_results_c))
    # print(f"Bootstrap CI estimates HLA-C: {round(x_low, 2)} {round(x_high, 2)}")
    # # DRB1
    # hlaupgrade_results_drb1: list[float] = [
    #     100 * feuille_hla_upgrade.cell(row=row_nu, column=30).value / 2 for row_nu in range(3, 2148)
    # ]
    # x_low, x_high = bootstrap_simulation(hlaupgrade_results_drb1, 25000, len(hlaupgrade_results_drb1))
    # print(f"Bootstrap CI estimates HLA-DRB1: {round(x_low, 2)} {round(x_high, 2)}")
    # # DQB1
    # hlaupgrade_results_dqb1: list[float] = [
    #     100 * feuille_hla_upgrade.cell(row=row_nu, column=32).value / 2 for row_nu in range(3, 2148)
    # ]
    # x_low, x_high = bootstrap_simulation(hlaupgrade_results_dqb1, 25000, len(hlaupgrade_results_dqb1))
    # print(f"Bootstrap CI estimates HLA-DQB1: {round(x_low, 2)} {round(x_high, 2)}")
    #
    # ## HLA-Upgrade (EU)
    # print("HLA-Upgrade (EU)")
    # document_hla_upgrade = openpyxl.load_workbook(
    #     r"C:\Users\lhott\Documents\Formation scolaire\These\Travail\SagaGL\HaploSFHI"
    #     r"\easyHLA Test Set\scoresEU.xlsx",
    #     data_only=True,
    # )
    # feuille_hla_upgrade = document_hla_upgrade["Feuil1"]
    #
    # # A
    # hlaupgrade_results_a: list[float] = [
    #     100 * feuille_hla_upgrade.cell(row=row_nu, column=23).value / 2 for row_nu in range(3, 2148)
    # ]
    # x_low, x_high = bootstrap_simulation(hlaupgrade_results_a, 25000, len(hlaupgrade_results_a))
    # print(f"Bootstrap CI estimates HLA-A: {round(x_low, 2)} {round(x_high, 2)}")
    # # B
    # hlaupgrade_results_b: list[float] = [
    #     100 * feuille_hla_upgrade.cell(row=row_nu, column=25).value / 2 for row_nu in range(3, 2148)
    # ]
    # x_low, x_high = bootstrap_simulation(hlaupgrade_results_b, 25000, len(hlaupgrade_results_b))
    # print(f"Bootstrap CI estimates HLA-B: {round(x_low, 2)} {round(x_high, 2)}")
    # # C
    # hlaupgrade_results_c: list[float] = [
    #     100 * feuille_hla_upgrade.cell(row=row_nu, column=27).value / 2 for row_nu in range(3, 2148)
    # ]
    # x_low, x_high = bootstrap_simulation(hlaupgrade_results_c, 25000, len(hlaupgrade_results_c))
    # print(f"Bootstrap CI estimates HLA-C: {round(x_low, 2)} {round(x_high, 2)}")
    # # DRB1
    # hlaupgrade_results_drb1: list[float] = [
    #     100 * feuille_hla_upgrade.cell(row=row_nu, column=29).value / 2 for row_nu in range(3, 2148)
    # ]
    # x_low, x_high = bootstrap_simulation(hlaupgrade_results_drb1, 25000, len(hlaupgrade_results_drb1))
    # print(f"Bootstrap CI estimates HLA-DRB1: {round(x_low, 2)} {round(x_high, 2)}")
    # # DQB1
    # hlaupgrade_results_dqb1: list[float] = [
    #     100 * feuille_hla_upgrade.cell(row=row_nu, column=31).value / 2 for row_nu in range(3, 2148)
    # ]
    # x_low, x_high = bootstrap_simulation(hlaupgrade_results_dqb1, 25000, len(hlaupgrade_results_dqb1))
    # print(f"Bootstrap CI estimates HLA-DQB1: {round(x_low, 2)} {round(x_high, 2)}")

    ### False Negative & False Positive Eplets
    # print("False Negative & False Positive Eplets")
    # for file, tool, fn_or_fp in zip(
    #     [
    #         r'C:\Users\lhott\Documents\Formation scolaire\These\Travail\SagaGL'
    #         r'\HaploSFHI\HaploStats Test Set\scoresCAU_pecc_fn_CL1.csv',
    #         r'C:\Users\lhott\Documents\Formation scolaire\These\Travail\SagaGL'
    #         r'\HaploSFHI\HaploStats Test Set\scoresCAU_pecc_fn_CL2.csv',
    #         r'C:\Users\lhott\Documents\Formation scolaire\These\Travail\SagaGL'
    #         r'\HaploSFHI\HaploStats Test Set\scoresCAU_pecc_fn_all.csv',
    #         r'C:\Users\lhott\Documents\Formation scolaire\These\Travail\SagaGL'
    #         r'\HaploSFHI\HaploStats Test Set\scoresCAU_pecc_fp_CL1.csv',
    #         r'C:\Users\lhott\Documents\Formation scolaire\These\Travail\SagaGL'
    #         r'\HaploSFHI\HaploStats Test Set\scoresCAU_pecc_fp_CL2.csv',
    #         r'C:\Users\lhott\Documents\Formation scolaire\These\Travail\SagaGL'
    #         r'\HaploSFHI\HaploStats Test Set\scoresCAU_pecc_fp_all.csv',
    #         r'C:\Users\lhott\Documents\Formation scolaire\These\Travail\SagaGL'
    #         r'\HaploSFHI\easyHLA Test Set\scoresEU_pecc_fn_CL1.csv',
    #         r'C:\Users\lhott\Documents\Formation scolaire\These\Travail\SagaGL'
    #         r'\HaploSFHI\easyHLA Test Set\scoresEU_pecc_fn_CL2.csv',
    #         r'C:\Users\lhott\Documents\Formation scolaire\These\Travail\SagaGL'
    #         r'\HaploSFHI\easyHLA Test Set\scoresEU_pecc_fn_all.csv',
    #         r'C:\Users\lhott\Documents\Formation scolaire\These\Travail\SagaGL'
    #         r'\HaploSFHI\easyHLA Test Set\scoresEU_pecc_fp_CL1.csv',
    #         r'C:\Users\lhott\Documents\Formation scolaire\These\Travail\SagaGL'
    #         r'\HaploSFHI\easyHLA Test Set\scoresEU_pecc_fp_CL2.csv',
    #         r'C:\Users\lhott\Documents\Formation scolaire\These\Travail\SagaGL'
    #         r'\HaploSFHI\easyHLA Test Set\scoresEU_pecc_fp_all.csv',
    #         r'C:\Users\lhott\Documents\Formation scolaire\These\Travail\SagaGL'
    #         r'\HaploSFHI\HaploSFHI Test Set\scoresSFHI_pecc_fn_CL1.csv',
    #         r'C:\Users\lhott\Documents\Formation scolaire\These\Travail\SagaGL'
    #         r'\HaploSFHI\HaploSFHI Test Set\scoresSFHI_pecc_fn_CL2.csv',
    #         r'C:\Users\lhott\Documents\Formation scolaire\These\Travail\SagaGL'
    #         r'\HaploSFHI\HaploSFHI Test Set\scoresSFHI_pecc_fn_all.csv',
    #         r'C:\Users\lhott\Documents\Formation scolaire\These\Travail\SagaGL'
    #         r'\HaploSFHI\HaploSFHI Test Set\scoresSFHI_pecc_fp_CL1.csv',
    #         r'C:\Users\lhott\Documents\Formation scolaire\These\Travail\SagaGL'
    #         r'\HaploSFHI\HaploSFHI Test Set\scoresSFHI_pecc_fp_CL2.csv',
    #         r'C:\Users\lhott\Documents\Formation scolaire\These\Travail\SagaGL'
    #         r'\HaploSFHI\HaploSFHI Test Set\scoresSFHI_pecc_fp_all.csv',
    #     ],
    #     [
    #         "HaploStats",
    #         "HaploStats",
    #         "HaploStats",
    #         "HaploStats",
    #         "HaploStats",
    #         "HaploStats",
    #         "HLA Upgrade",
    #         "HLA Upgrade",
    #         "HLA Upgrade",
    #         "HLA Upgrade",
    #         "HLA Upgrade",
    #         "HLA Upgrade",
    #         "HaploSFHI",
    #         "HaploSFHI",
    #         "HaploSFHI",
    #         "HaploSFHI",
    #         "HaploSFHI",
    #         "HaploSFHI",
    #     ],
    #     [
    #         "FN",
    #         "FN",
    #         "FN",
    #         "FP",
    #         "FP",
    #         "FP",
    #         "FN",
    #         "FN",
    #         "FN",
    #         "FP",
    #         "FP",
    #         "FP",
    #         "FN",
    #         "FN",
    #         "FN",
    #         "FP",
    #         "FP",
    #         "FP",
    #     ]
    #
    # ):
    #     with open(file, 'r') as csvfile:
    #         csvreader = csv.reader(csvfile)
    #         # skip first row
    #         next(csvreader)
    #         haplostats_results: list[float] = [float(row[1]) for row in csvreader]
    #         continue_: bool = True
    #         x_low, x_high = bootstrap_simulation(haplostats_results, 25000, len(haplostats_results))
    #         mean_ = sum(haplostats_results) / len(haplostats_results)
    #         print(f"Bootstrap CI estimates {tool} {fn_or_fp}:"
    #               f" {round(x_low, 2)} {round(mean_, 2)} {round(x_high, 2)}")


    ## HAPLOSFHI SECOND PAPER TEST DATASET ###
    # HaploStats
    feuille_haplostats = openpyxl.load_workbook(
        r"C:\Users\lhott\Documents\Formation scolaire\These\Travail"
        r"\HLATypingExtrapolation\OtherTools\HLA Upgrade\scores_easyHLA.xlsx"
    ).worksheets[3]
    # A
    haplostats_cau_results_a: list[float] = [
        100 * feuille_haplostats.cell(row=row_nu, column=31).value / 2 for row_nu in range(3, 7699)
    ]
    assert len(haplostats_cau_results_a) == 7696
    x_low, x_high = bootstrap_simulation(haplostats_cau_results_a, 25000, len(haplostats_cau_results_a))
    print(rf"Bootstrap CI estimates HLA-A: \\ {{[}}{round(x_low, 2)}-{round(x_high, 2)}{{]}}")
    # B
    haplostats_cau_results_b: list[float] = [
        100 * feuille_haplostats.cell(row=row_nu, column=33).value / 2 for row_nu in range(3, 7699)
    ]
    assert len(haplostats_cau_results_b) == 7696
    x_low, x_high = bootstrap_simulation(haplostats_cau_results_b, 25000, len(haplostats_cau_results_b))
    print(rf"Bootstrap CI estimates HLA-B: \\ {{[}}{round(x_low, 2)}-{round(x_high, 2)}{{]}}")
    # C
    haplostats_cau_results_c: list[float] = [
        100 * feuille_haplostats.cell(row=row_nu, column=35).value / 2 for row_nu in range(3, 7699)
    ]
    assert len(haplostats_cau_results_c) == 7696
    x_low, x_high = bootstrap_simulation(haplostats_cau_results_c, 25000, len(haplostats_cau_results_c))
    print(rf"Bootstrap CI estimates HLA-C: \\ {{[}}{round(x_low, 2)}-{round(x_high, 2)}{{]}}")
    # DRB1
    haplostats_cau_results_drb1: list[float] = [
        100 * feuille_haplostats.cell(row=row_nu, column=37).value / 2 for row_nu in range(3, 7699)
    ]
    assert len(haplostats_cau_results_drb1) == 7696
    x_low, x_high = bootstrap_simulation(haplostats_cau_results_drb1, 25000, len(haplostats_cau_results_drb1))
    print(rf"Bootstrap CI estimates HLA-DRB1: \\ {{[}}{round(x_low, 2)}-{round(x_high, 2)}{{]}}")
    # DQB1
    haplostats_cau_results_dqb1: list[float] = [
        100 * feuille_haplostats.cell(row=row_nu, column=39).value / 2 for row_nu in range(3, 7699)
    ]
    assert len(haplostats_cau_results_dqb1) == 7696
    x_low, x_high = bootstrap_simulation(haplostats_cau_results_dqb1, 25000, len(haplostats_cau_results_dqb1))
    print(rf"Bootstrap CI estimates HLA-DQB1: \\ {{[}}{round(x_low, 2)}-{round(x_high, 2)}{{]}}")
    # # DRB345
    # haplostats_cau_results_drb345: list[float] = [
    #     100 * feuille_haplostats.cell(row=row_nu, column=43).value / 2 for row_nu in range(3, 7699)
    # ]
    # assert len(haplostats_cau_results_drb345) == 7696
    # x_low, x_high = bootstrap_simulation(haplostats_cau_results_drb345, 25000, len(haplostats_cau_results_drb345))
    # print(rf"Bootstrap CI estimates HLA-DRB345: \\ {{[}}{round(x_low, 2)}-{round(x_high, 2)}{{]}}")

    # # HaploSFHI
    # feuille_haplosfhi = openpyxl.load_workbook(
    #     r"C:\Users\lhott\Documents\Formation scolaire\These\Travail"
    #     r"\HLATypingExtrapolation\OtherTools\HLA EMMA\scores_EMMA_NL.xlsx"
    # ).worksheets[1]
    # feuille_haplosfhi_dp = openpyxl.load_workbook(
    #     r"C:\Users\lhott\Documents\Formation scolaire\These\Travail"
    #     r"\HLATypingExtrapolation\OtherTools\HLA EMMA\scores_EMMA_NL.xlsx"
    # ).worksheets[3]
    # # A
    # feuille_haplosfhi_results_a: list[float] = [
    #     100 * feuille_haplosfhi.cell(row=row_nu, column=31).value / 2 for row_nu in range(3, 7699)
    # ]
    # assert len(feuille_haplosfhi_results_a) == 7696
    # x_low, x_high = bootstrap_simulation(feuille_haplosfhi_results_a, 25000, len(feuille_haplosfhi_results_a))
    # print(rf"Bootstrap CI estimates HLA-A: \\ \tiny{{{{[}}{round(x_low, 2):.2f}-{round(x_high, 2):.2f}{{]}}}}")
    # # B
    # feuille_haplosfhi_results_b: list[float] = [
    #     100 * feuille_haplosfhi.cell(row=row_nu, column=33).value / 2 for row_nu in range(3, 7699)
    # ]
    # assert len(feuille_haplosfhi_results_b) == 7696
    # x_low, x_high = bootstrap_simulation(feuille_haplosfhi_results_b, 25000, len(feuille_haplosfhi_results_b))
    # print(rf"Bootstrap CI estimates HLA-B: \\ \tiny{{{{[}}{round(x_low, 2):.2f}-{round(x_high, 2):.2f}{{]}}}}")
    # # C
    # feuille_haplosfhi_results_c: list[float] = [
    #     100 * feuille_haplosfhi.cell(row=row_nu, column=35).value / 2 for row_nu in range(3, 7699)
    # ]
    # assert len(feuille_haplosfhi_results_c) == 7696
    # x_low, x_high = bootstrap_simulation(feuille_haplosfhi_results_c, 25000, len(feuille_haplosfhi_results_c))
    # print(rf"Bootstrap CI estimates HLA-C: \\ \tiny{{{{[}}{round(x_low, 2):.2f}-{round(x_high, 2):.2f}{{]}}}}")
    # # DRB1
    # feuille_haplosfhi_results_drb1: list[float] = [
    #     100 * feuille_haplosfhi.cell(row=row_nu, column=37).value / 2 for row_nu in range(3, 7699)
    # ]
    # assert len(feuille_haplosfhi_results_drb1) == 7696
    # x_low, x_high = bootstrap_simulation(feuille_haplosfhi_results_drb1, 25000, len(feuille_haplosfhi_results_drb1))
    # print(rf"Bootstrap CI estimates HLA-DRB1: \\ \tiny{{{{[}}{round(x_low, 2):.2f}-{round(x_high, 2):.2f}{{]}}}}")
    # # DQB1
    # feuille_haplosfhi_results_dqb1: list[float] = [
    #     100 * feuille_haplosfhi.cell(row=row_nu, column=39).value / 2 for row_nu in range(3, 7699)
    # ]
    # assert len(feuille_haplosfhi_results_dqb1) == 7696
    # x_low, x_high = bootstrap_simulation(feuille_haplosfhi_results_dqb1, 25000, len(feuille_haplosfhi_results_dqb1))
    # print(rf"Bootstrap CI estimates HLA-DQB1: \\ \tiny{{{{[}}{round(x_low, 2):.2f}-{round(x_high, 2):.2f}{{]}}}}")
    # # DQA1
    # feuille_haplosfhi_results_dqa1: list[float] = [
    #     100 * feuille_haplosfhi.cell(row=row_nu, column=41).value / 2 for row_nu in range(3, 7699)
    # ]
    # assert len(feuille_haplosfhi_results_dqa1) == 7696
    # x_low, x_high = bootstrap_simulation(feuille_haplosfhi_results_dqa1, 25000, len(feuille_haplosfhi_results_dqa1))
    # print(rf"Bootstrap CI estimates HLA-DQA1: \\ \tiny{{{{[}}{round(x_low, 2):.2f}-{round(x_high, 2):.2f}{{]}}}}")
    # # DRB345
    # feuille_haplosfhi_results_drb345: list[float] = [
    #     100 * feuille_haplosfhi.cell(row=row_nu, column=43).value / 2 for row_nu in range(3, 7699)
    # ]
    # assert len(feuille_haplosfhi_results_drb345) == 7696
    # x_low, x_high = bootstrap_simulation(feuille_haplosfhi_results_drb345, 25000, len(feuille_haplosfhi_results_drb345))
    # print(rf"Bootstrap CI estimates HLA-DRB345: \\ \tiny{{{{[}}{round(x_low, 2):.2f}-{round(x_high, 2):.2f}{{]}}}}")
    # DPB1
    # feuille_haplosfhi_results_dpb1: list[float] = [
    #     100 * feuille_haplosfhi_dp.cell(row=row_nu, column=11).value / 2 for row_nu in range(3, 7699)
    # ]
    # assert len(feuille_haplosfhi_results_dpb1) == 7696
    # x_low, x_high = bootstrap_simulation(feuille_haplosfhi_results_dpb1, 25000, len(feuille_haplosfhi_results_dpb1))
    # print(rf"Bootstrap CI estimates HLA-DPB1: \\ \tiny{{{{[}}{round(x_low, 2):.2f}-{round(x_high, 2):.2f}{{]}}}}")
    # # DPA1
    # feuille_haplosfhi_results_dpa1: list[float] = [
    #     100 * feuille_haplosfhi_dp.cell(row=row_nu, column=13).value / 2 for row_nu in range(3, 7699)
    # ]
    # assert len(feuille_haplosfhi_results_dpa1) == 7696
    # x_low, x_high = bootstrap_simulation(feuille_haplosfhi_results_dpa1, 25000, len(feuille_haplosfhi_results_dpa1))
    # print(rf"Bootstrap CI estimates HLA-DPA1: \\ \tiny{{{{[}}{round(x_low, 2):.2f}-{round(x_high, 2):.2f}{{]}}}}")


    ## FOREIGN DATASETS ###
    # # Russia (all tools)
    # feuille_haplostats = openpyxl.load_workbook(
    #     r"C:\Users\lhott\Documents\Formation scolaire\These\Travail"
    #     r"\data_allelefrequencies\Testing Tools On Foreign Data\Russia\HLA EMMA\scoresEMMA_NL.xlsx"
    # ).worksheets[1]
    # # A
    # haplostats_cau_results_a: list[float] = [
    #     100 * feuille_haplostats.cell(row=row_nu, column=23).value / 2 for row_nu in range(3, 1513)
    # ]
    # assert len(haplostats_cau_results_a) == 1510
    # x_low, x_high = bootstrap_simulation(haplostats_cau_results_a, 25000, len(haplostats_cau_results_a))
    # print(rf"Bootstrap CI estimates HLA-A: \\ \tiny{{{{[}}{round(x_low, 2):.2f}-{round(x_high, 2):.2f}{{]}}}}")
    # # B
    # haplostats_cau_results_b: list[float] = [
    #     100 * feuille_haplostats.cell(row=row_nu, column=25).value / 2 for row_nu in range(3, 1513)
    # ]
    # assert len(haplostats_cau_results_b) == 1510
    # x_low, x_high = bootstrap_simulation(haplostats_cau_results_b, 25000, len(haplostats_cau_results_b))
    # print(rf"Bootstrap CI estimates HLA-B: \\ \tiny{{{{[}}{round(x_low, 2):.2f}-{round(x_high, 2):.2f}{{]}}}}")
    # # C
    # haplostats_cau_results_c: list[float] = [
    #     100 * feuille_haplostats.cell(row=row_nu, column=27).value / 2 for row_nu in range(3, 1513)
    # ]
    # assert len(haplostats_cau_results_c) == 1510
    # x_low, x_high = bootstrap_simulation(haplostats_cau_results_c, 25000, len(haplostats_cau_results_c))
    # print(rf"Bootstrap CI estimates HLA-C: \\ \tiny{{{{[}}{round(x_low, 2):.2f}-{round(x_high, 2):.2f}{{]}}}}")
    # # DRB1
    # haplostats_cau_results_drb1: list[float] = [
    #     100 * feuille_haplostats.cell(row=row_nu, column=29).value / 2 for row_nu in range(3, 1513)
    # ]
    # assert len(haplostats_cau_results_drb1) == 1510
    # x_low, x_high = bootstrap_simulation(haplostats_cau_results_drb1, 25000, len(haplostats_cau_results_drb1))
    # print(rf"Bootstrap CI estimates HLA-DRB1: \\ \tiny{{{{[}}{round(x_low, 2):.2f}-{round(x_high, 2):.2f}{{]}}}}")
    # # DQB1
    # haplostats_cau_results_dqb1: list[float] = [
    #     100 * feuille_haplostats.cell(row=row_nu, column=31).value / 2 for row_nu in range(3, 1513)
    # ]
    # assert len(haplostats_cau_results_dqb1) == 1510
    # x_low, x_high = bootstrap_simulation(haplostats_cau_results_dqb1, 25000, len(haplostats_cau_results_dqb1))
    # print(rf"Bootstrap CI estimates HLA-DQB1: \\ \tiny{{{{[}}{round(x_low, 2):.2f}-{round(x_high, 2):.2f}{{]}}}}")

    # ### Inter -> High resolution (all tools) ###
    # feuille_haplostats = openpyxl.load_workbook(
    #     r"C:\Users\lhott\Documents\Formation scolaire\These\Travail"
    #     r"\inter_resolution_imputation\HaploStats\scoresEURCAU.xlsx"
    # ).worksheets[1]
    # # A
    # haplostats_cau_results_a: list[float] = [
    #     100 * feuille_haplostats.cell(row=row_nu, column=23).value / 2 for row_nu in range(3, 360)
    # ]
    # assert len(haplostats_cau_results_a) == 357
    # x_low, x_high = bootstrap_simulation(haplostats_cau_results_a, 25000, len(haplostats_cau_results_a))
    # print(rf"Bootstrap CI estimates HLA-A: \\ \tiny{{{{[}}{round(x_low, 2)}-{round(x_high, 2)}{{]}}}}")
    # # B
    # haplostats_cau_results_b: list[float] = [
    #     100 * feuille_haplostats.cell(row=row_nu, column=25).value / 2 for row_nu in range(3, 360)
    # ]
    # assert len(haplostats_cau_results_b) == 357
    # x_low, x_high = bootstrap_simulation(haplostats_cau_results_b, 25000, len(haplostats_cau_results_b))
    # print(rf"Bootstrap CI estimates HLA-B: \\ \tiny{{{{[}}{round(x_low, 2)}-{round(x_high, 2)}{{]}}}}")
    # # C
    # haplostats_cau_results_c: list[float] = [
    #     100 * feuille_haplostats.cell(row=row_nu, column=27).value / 2 for row_nu in range(3, 360)
    # ]
    # assert len(haplostats_cau_results_c) == 357
    # x_low, x_high = bootstrap_simulation(haplostats_cau_results_c, 25000, len(haplostats_cau_results_c))
    # print(rf"Bootstrap CI estimates HLA-C: \\ \tiny{{{{[}}{round(x_low, 2)}-{round(x_high, 2)}{{]}}}}")
    # # DRB1
    # haplostats_cau_results_drb1: list[float] = [
    #     100 * feuille_haplostats.cell(row=row_nu, column=29).value / 2 for row_nu in range(3, 360)
    # ]
    # assert len(haplostats_cau_results_drb1) == 357
    # x_low, x_high = bootstrap_simulation(haplostats_cau_results_drb1, 25000, len(haplostats_cau_results_drb1))
    # print(rf"Bootstrap CI estimates HLA-DRB1: \\ \tiny{{{{[}}{round(x_low, 2)}-{round(x_high, 2)}{{]}}}}")
    # # DQB1
    # haplostats_cau_results_dqb1: list[float] = [
    #     100 * feuille_haplostats.cell(row=row_nu, column=31).value / 2 for row_nu in range(3, 360)
    # ]
    # assert len(haplostats_cau_results_dqb1) == 357
    # x_low, x_high = bootstrap_simulation(haplostats_cau_results_dqb1, 25000, len(haplostats_cau_results_dqb1))
    # print(rf"Bootstrap CI estimates HLA-DQB1: \\ \tiny{{{{[}}{round(x_low, 2)}-{round(x_high, 2)}{{]}}}}")

    # ## London (all tools) ###
    # feuille_haplostats = openpyxl.load_workbook(
    #     r"C:\Users\lhott\Documents\Formation scolaire\These\Travail"
    #     r"\extra_foreign_ngs_data\london\WhiteCaucasian\HLA EMMA\scoresEMMA_NL.xlsx"
    # ).worksheets[1]
    #
    # sample_size = 361
    # last_line = 364
    # # A
    # haplostats_cau_results_a: list[float] = [
    #     100 * feuille_haplostats.cell(row=row_nu, column=23).value / 2 for row_nu in range(3, last_line)
    # ]
    # assert len(haplostats_cau_results_a) == sample_size
    # x_low, x_high = bootstrap_simulation(haplostats_cau_results_a, 25000, len(haplostats_cau_results_a))
    # print(rf"Bootstrap CI estimates HLA-A: \\ \tiny{{{{[}}{round(x_low, 2):.2f}-{round(x_high, 2):.2f}{{]}}}}")
    # # B
    # haplostats_cau_results_b: list[float] = [
    #     100 * feuille_haplostats.cell(row=row_nu, column=25).value / 2 for row_nu in range(3, last_line)
    # ]
    # assert len(haplostats_cau_results_b) == sample_size
    # x_low, x_high = bootstrap_simulation(haplostats_cau_results_b, 25000, len(haplostats_cau_results_b))
    # print(rf"Bootstrap CI estimates HLA-B: \\ \tiny{{{{[}}{round(x_low, 2):.2f}-{round(x_high, 2):.2f}{{]}}}}")
    # # C
    # haplostats_cau_results_c: list[float] = [
    #     100 * feuille_haplostats.cell(row=row_nu, column=27).value / 2 for row_nu in range(3, last_line)
    # ]
    # assert len(haplostats_cau_results_c) == sample_size
    # x_low, x_high = bootstrap_simulation(haplostats_cau_results_c, 25000, len(haplostats_cau_results_c))
    # print(rf"Bootstrap CI estimates HLA-C: \\ \tiny{{{{[}}{round(x_low, 2):.2f}-{round(x_high, 2):.2f}{{]}}}}")
    # # DRB1
    # haplostats_cau_results_drb1: list[float] = [
    #     100 * feuille_haplostats.cell(row=row_nu, column=29).value / 2 for row_nu in range(3, last_line)
    # ]
    # assert len(haplostats_cau_results_drb1) == sample_size
    # x_low, x_high = bootstrap_simulation(haplostats_cau_results_drb1, 25000, len(haplostats_cau_results_drb1))
    # print(rf"Bootstrap CI estimates HLA-DRB1: \\ \tiny{{{{[}}{round(x_low, 2):.2f}-{round(x_high, 2):.2f}{{]}}}}")
    # # DQB1
    # haplostats_cau_results_dqb1: list[float] = [
    #     100 * feuille_haplostats.cell(row=row_nu, column=31).value / 2 for row_nu in range(3, last_line)
    # ]
    # assert len(haplostats_cau_results_dqb1) == sample_size
    # x_low, x_high = bootstrap_simulation(haplostats_cau_results_dqb1, 25000, len(haplostats_cau_results_dqb1))
    # print(rf"Bootstrap CI estimates HLA-DQB1: \\ \tiny{{{{[}}{round(x_low, 2):.2f}-{round(x_high, 2):.2f}{{]}}}}")

    # ### Porto / Scotland (all tools) ###
    # feuille_haplostats = openpyxl.load_workbook(
    #     r"C:\Users\lhott\Documents\Formation scolaire\These\Travail"
    #     r"\extra_foreign_ngs_data\porto\HLA EMMA\scoresEMMA_NL.xlsx"
    # ).worksheets[1]
    #
    # sample_size = 215
    # last_line = 218
    # # A
    # haplostats_cau_results_a: list[float] = [
    #     100 * feuille_haplostats.cell(row=row_nu, column=27).value / 2 for row_nu in range(3, last_line)
    # ]
    # assert len(haplostats_cau_results_a) == sample_size
    # x_low, x_high = bootstrap_simulation(haplostats_cau_results_a, 25000, len(haplostats_cau_results_a))
    # print(rf"Bootstrap CI estimates HLA-A: \\ \tiny{{{{[}}{round(x_low, 2):.2f}-{round(x_high, 2):.2f}{{]}}}}")
    # # B
    # haplostats_cau_results_b: list[float] = [
    #     100 * feuille_haplostats.cell(row=row_nu, column=29).value / 2 for row_nu in range(3, last_line)
    # ]
    # assert len(haplostats_cau_results_b) == sample_size
    # x_low, x_high = bootstrap_simulation(haplostats_cau_results_b, 25000, len(haplostats_cau_results_b))
    # print(rf"Bootstrap CI estimates HLA-B: \\ \tiny{{{{[}}{round(x_low, 2):.2f}-{round(x_high, 2):.2f}{{]}}}}")
    # # C
    # haplostats_cau_results_c: list[float] = [
    #     100 * feuille_haplostats.cell(row=row_nu, column=31).value / 2 for row_nu in range(3, last_line)
    # ]
    # assert len(haplostats_cau_results_c) == sample_size
    # x_low, x_high = bootstrap_simulation(haplostats_cau_results_c, 25000, len(haplostats_cau_results_c))
    # print(rf"Bootstrap CI estimates HLA-C: \\ \tiny{{{{[}}{round(x_low, 2):.2f}-{round(x_high, 2):.2f}{{]}}}}")
    # # DRB1
    # haplostats_cau_results_drb1: list[float] = [
    #     100 * feuille_haplostats.cell(row=row_nu, column=33).value / 2 for row_nu in range(3, last_line)
    # ]
    # assert len(haplostats_cau_results_drb1) == sample_size
    # x_low, x_high = bootstrap_simulation(haplostats_cau_results_drb1, 25000, len(haplostats_cau_results_drb1))
    # print(rf"Bootstrap CI estimates HLA-DRB1: \\ \tiny{{{{[}}{round(x_low, 2):.2f}-{round(x_high, 2):.2f}{{]}}}}")
    # # DQB1
    # haplostats_cau_results_dqb1: list[float] = [
    #     100 * feuille_haplostats.cell(row=row_nu, column=35).value / 2 for row_nu in range(3, last_line)
    # ]
    # assert len(haplostats_cau_results_dqb1) == sample_size
    # x_low, x_high = bootstrap_simulation(haplostats_cau_results_dqb1, 25000, len(haplostats_cau_results_dqb1))
    # print(rf"Bootstrap CI estimates HLA-DQB1: \\ \tiny{{{{[}}{round(x_low, 2):.2f}-{round(x_high, 2):.2f}{{]}}}}")
    # # DQA1
    # haplostats_cau_results_dqa1: list[float] = [
    #     100 * feuille_haplostats.cell(row=row_nu, column=37).value / 2 for row_nu in range(3, last_line)
    # ]
    # assert len(haplostats_cau_results_dqa1) == sample_size
    # x_low, x_high = bootstrap_simulation(haplostats_cau_results_dqa1, 25000, len(haplostats_cau_results_dqa1))
    # print(rf"Bootstrap CI estimates HLA-DQA1: \\ \tiny{{{{[}}{round(x_low, 2):.2f}-{round(x_high, 2):.2f}{{]}}}}")


if __name__ == "__main__":
    haplosfhi_bootstrap_simulation()
